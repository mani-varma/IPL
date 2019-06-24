import os
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'ipl.settings')
import django
django.setup()

import openpyxl
from iplapp.models import *

def importMatchesData():
    book = openpyxl.load_workbook("matches.xlsx")
    sheet = book.get_active_sheet()
    for i in range(2, sheet.max_row+1):
        match = Matches(
        mid=sheet.cell(row=i, column=1).value,
        season = sheet.cell(row=i, column=2).value,
        city = sheet.cell(row=i, column=3).value,
        date = sheet.cell(row=i, column=4).value,
        team1 = sheet.cell(row=i, column=5).value,
        team2 = sheet.cell(row=i, column=6).value,
        toss_winner = sheet.cell(row=i, column=7).value,
        toss_decision = sheet.cell(row=i, column=8).value,
        result = sheet.cell(row=i, column=9).value,
        dl_applied = sheet.cell(row=i, column=10).value,
        winner = sheet.cell(row=i, column=11).value,
        win_by_runs = sheet.cell(row=i, column=12).value,
        win_by_wickets = sheet.cell(row=i, column=13).value,
        player_of_match = sheet.cell(row=i, column=14).value,
        venue = sheet.cell(row=i, column=15).value,
        umpire1 = sheet.cell(row=i, column=16).value,
        umpire2 = sheet.cell(row=i, column=17).value,
        umpire3 = sheet.cell(row=i, column=18).value,
        )
        match.save()
    print("Data imported Succesfully.")


def importDeliveriesData():
    print("HI1")
    book = openpyxl.load_workbook("deliveries.xlsx")
    print("HI2")
    sheet = book.get_active_sheet()
    print("HI3")
    for i in range(2, sheet.max_row + 1):
        print(sheet.cell(row=i, column=1).value)
        delivery=Deliveries(
        mdid=sheet.cell(row=i, column=1).value,
        innings = sheet.cell(row=i, column=2).value,
        batting_team = sheet.cell(row=i, column=3).value,
        bowling_team = sheet.cell(row=i, column=4).value,
        over = sheet.cell(row=i, column=5).value,
        ball = sheet.cell(row=i, column=6).value,
        batsman = sheet.cell(row=i, column=7).value,
        non_striker = sheet.cell(row=i, column=8).value,
        bowler = sheet.cell(row=i, column=9).value,
        is_super_over = sheet.cell(row=i, column=10).value,
        wide_runs = sheet.cell(row=i, column=11).value,
        bye_runs = sheet.cell(row=i, column=12).value,
        legbye_runs = sheet.cell(row=i, column=13).value,
        noball_runs = sheet.cell(row=i, column=14).value,
        penalty_runs = sheet.cell(row=i, column=15).value,
        batsman_runs = sheet.cell(row=i, column=16).value,
        extra_runs = sheet.cell(row=i, column=17).value,
        total_runs = sheet.cell(row=i, column=18).value,
        player_dismissed = sheet.cell(row=i, column=19).value,
        dismissal_kind = sheet.cell(row=i, column=20).value,
        fielder = sheet.cell(row=i, column=21).value,
        match=Matches.objects.filter(mid=sheet.cell(row=i, column=1).value)[0],
        )
        delivery.save()


if __name__ == '__main__':
    #importMatchesData()
    importDeliveriesData()